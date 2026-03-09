#!/usr/bin/env python3
"""
Archive Parity Review Spreadsheet Generator
Compares legacy FYD website pages with new site pages and generates a migration review spreadsheet.
"""

import os
import re
from pathlib import Path
from html.parser import HTMLParser
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    DEFAULT_FONT
)
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION
# ============================================================================

OLD_SITE_PATH = "/sessions/trusting-elegant-johnson/mnt/web/OldFYDWebSite/wwwroot/wwwroot"
NEW_SITE_PATH = "/sessions/trusting-elegant-johnson/mnt/web/NewFYDWebSite/www"
OUTPUT_PATH = "/sessions/trusting-elegant-johnson/mnt/web/NewFYDWebSite/www/archive/FarrDesigns_Parity_Review.xlsx"

# Color definitions (RGB)
COLOR_KEEP = "C6EFCE"      # Light green
COLOR_REDIRECT = "FFEB9C"  # Light yellow
COLOR_DEFER = "D9D9D9"     # Light gray
COLOR_RETIRE = "FFC7CE"    # Light red
COLOR_HEADER = "4472C4"    # Blue


# ============================================================================
# HTML TITLE EXTRACTOR
# ============================================================================

class TitleExtractor(HTMLParser):
    """Extract title tag from HTML"""
    def __init__(self):
        super().__init__()
        self.title = None
        self.in_title = False

    def handle_starttag(self, tag, attrs):
        if tag.lower() == 'title':
            self.in_title = True

    def handle_endtag(self, tag):
        if tag.lower() == 'title':
            self.in_title = False

    def handle_data(self, data):
        if self.in_title and not self.title:
            self.title = data.strip()


def extract_title(html_path):
    """Extract title from an HTML file"""
    try:
        with open(html_path, 'r', encoding='utf-8', errors='ignore') as f:
            parser = TitleExtractor()
            # Only parse first 2KB to get the title quickly
            content = f.read(2000)
            parser.feed(content)
            return parser.title or ""
    except Exception as e:
        print(f"Error reading {html_path}: {e}")
        return ""


def extract_design_number(filename):
    """Extract design number from filename (e.g., '100.html' -> '100')"""
    # Match patterns like 100.html, 100m.html, 028results.html
    match = re.match(r'^(\d+)', filename)
    if match:
        return match.group(1)
    return None


def extract_boat_name(title):
    """Extract boat name from title
    Examples:
    'FYD | Farr 44 (Design 100)' -> 'Farr 44'
    '10' General Purpose Dinghy — Farr Yacht Design' -> '10\' General Purpose Dinghy'
    """
    if not title:
        return None

    # Modern yacht page format: "Boat Name — Farr Yacht Design"
    if "—" in title:
        parts = title.split("—")
        return parts[0].strip()

    # Legacy format: "FYD | Boat Name (Design XXX)"
    if "|" in title:
        parts = title.split("|")
        if len(parts) > 1:
            content = parts[1].strip()
            # Extract name before the design number
            match = re.match(r'^(.+?)\s*\(Design', content)
            if match:
                return match.group(1).strip()
            return content

    return None


def is_legacy_stub(html_path):
    """Determine if a new site page is a legacy stub (unchanged from old site)

    Heuristics:
    - Contains "adobe" or "muse" (Adobe Muse markers)
    - Has css files with crc values (Adobe Muse pattern)
    - Contains "museutils.js" or "museconfig.js"
    - No modern nav structure (doesn't have '../portfolio.html' or 'nav-logo')
    """
    try:
        with open(html_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()

        # Check for Adobe Muse markers
        if "museutils.js" in content or "museconfig.js" in content:
            return True
        if "generator" in content and "Muse" in content:
            return True
        if 'crc=' in content:  # Muse CSS pattern
            return True

        # Check for modern nav structure
        if "nav-logo" in content or "../portfolio.html" in content:
            return False

        return False
    except:
        return False


# ============================================================================
# SITE SCANNING
# ============================================================================

def scan_old_site():
    """Scan legacy site and extract metadata"""
    pages = {}
    old_path = Path(OLD_SITE_PATH)

    print(f"Scanning legacy site: {OLD_SITE_PATH}")
    for html_file in sorted(old_path.glob("*.html")):
        filename = html_file.name
        title = extract_title(str(html_file))
        design_num = extract_design_number(filename)
        boat_name = extract_boat_name(title)

        pages[filename] = {
            "filename": filename,
            "design_number": design_num,
            "boat_name": boat_name,
            "title": title,
        }

    print(f"Found {len(pages)} pages in legacy site")
    return pages


def scan_new_site():
    """Scan new site and extract metadata"""
    pages = {}
    new_path = Path(NEW_SITE_PATH)

    print(f"Scanning new site: {NEW_SITE_PATH}")

    # Scan root HTML files (numbered and named pages)
    for html_file in sorted(new_path.glob("*.html")):
        filename = html_file.name
        title = extract_title(str(html_file))
        design_num = extract_design_number(filename)
        boat_name = extract_boat_name(title)
        is_stub = is_legacy_stub(str(html_file))

        status = "Legacy stub" if is_stub else "Modern page"

        pages[filename] = {
            "filename": filename,
            "path": f"/{filename}",
            "design_number": design_num,
            "boat_name": boat_name,
            "title": title,
            "status": status,
            "is_stub": is_stub,
        }

    # Scan yacht subdirectory (modern pages)
    yacht_path = new_path / "yacht"
    if yacht_path.exists():
        for html_file in sorted(yacht_path.glob("*.html")):
            filename = html_file.name
            title = extract_title(str(html_file))
            boat_name = extract_boat_name(title)

            pages[filename] = {
                "filename": filename,
                "path": f"/yacht/{filename}",
                "design_number": None,  # yacht pages don't have design numbers
                "boat_name": boat_name,
                "title": title,
                "status": "Modern page",
                "is_stub": False,
            }

    print(f"Found {len(pages)} pages in new site")
    return pages


# ============================================================================
# MATCHING AND CLASSIFICATION
# ============================================================================

def find_matching_new_page(legacy_filename, legacy_design, old_pages, new_pages):
    """Find matching new site page for a legacy page

    Strategy:
    1. If it's a numbered page, look for corresponding numbered page in new site
    2. If there's a yacht subpage with matching design number, note that
    3. Otherwise mark as "Not migrated"
    """
    design_num = extract_design_number(legacy_filename)

    # First check: exact filename match in new site root
    if legacy_filename in new_pages:
        return legacy_filename, "root"

    # Second check: look for yacht page with design number
    if design_num:
        for new_filename, new_page in new_pages.items():
            if new_page.get("path", "").startswith("/yacht/"):
                # Check if yacht page title contains the design number
                if design_num in new_page.get("title", ""):
                    return new_filename, "yacht"

    return None, None


def classify_page(legacy_filename, legacy_data, matching_new_file, new_pages):
    """Classify a page and determine the decision

    Decision logic:
    - "Keep (modern)": Matching page in new site is modern (not a stub)
    - "Redirect": Matching page exists but is a legacy stub
    - "Defer": No modern equivalent, needs future attention
    - "Retire": Results pages, photo galleries, etc.
    """
    # Identify pages to retire
    if "results" in legacy_filename.lower():
        return "Retire", "Results/competition page"
    if "photo" in legacy_filename.lower():
        return "Retire", "Photo gallery"
    if "gallery" in legacy_filename.lower():
        return "Retire", "Photo gallery"

    if matching_new_file:
        new_page = new_pages.get(matching_new_file, {})
        if new_page.get("is_stub"):
            return "Redirect", "Legacy stub - redirect to modern equivalent"
        else:
            return "Keep (modern)", "Modern page with real content"
    else:
        return "Defer", "Legacy stub with no modern equivalent - review needed"


def generate_review_data(old_pages, new_pages):
    """Generate review data by matching and classifying all pages"""
    review_rows = []

    for legacy_filename, legacy_data in sorted(old_pages.items()):
        matching_file, match_type = find_matching_new_page(
            legacy_filename,
            legacy_data.get("design_number"),
            old_pages,
            new_pages
        )

        if matching_file and matching_file in new_pages:
            new_page_data = new_pages[matching_file]
            new_site_page = new_page_data["path"]
            new_site_status = new_page_data.get("status", "Unknown")
        else:
            new_site_page = None
            new_site_status = "Not migrated"

        decision, notes = classify_page(
            legacy_filename,
            legacy_data,
            matching_file,
            new_pages
        )

        review_rows.append({
            "legacy_page": legacy_filename,
            "design_number": legacy_data.get("design_number") or "",
            "boat_name": legacy_data.get("boat_name") or "",
            "new_site_page": new_site_page or "",
            "new_site_status": new_site_status,
            "decision": decision,
            "notes": notes,
        })

    return review_rows


# ============================================================================
# SPREADSHEET GENERATION
# ============================================================================

def create_spreadsheet(review_rows):
    """Create Excel spreadsheet with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Parity Review"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Define color fills for decisions
    decision_colors = {
        "Keep (modern)": PatternFill(start_color=COLOR_KEEP, end_color=COLOR_KEEP, fill_type="solid"),
        "Redirect": PatternFill(start_color=COLOR_REDIRECT, end_color=COLOR_REDIRECT, fill_type="solid"),
        "Defer": PatternFill(start_color=COLOR_DEFER, end_color=COLOR_DEFER, fill_type="solid"),
        "Retire": PatternFill(start_color=COLOR_RETIRE, end_color=COLOR_RETIRE, fill_type="solid"),
    }

    # Headers
    headers = [
        "Legacy Page",
        "Design Number",
        "Boat Name",
        "New Site Page",
        "New Site Status",
        "Decision",
        "Notes"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Data rows
    for row_num, row_data in enumerate(review_rows, 2):
        ws.cell(row=row_num, column=1).value = row_data["legacy_page"]
        ws.cell(row=row_num, column=2).value = row_data["design_number"]
        ws.cell(row=row_num, column=3).value = row_data["boat_name"]
        ws.cell(row=row_num, column=4).value = row_data["new_site_page"]
        ws.cell(row=row_num, column=5).value = row_data["new_site_status"]
        ws.cell(row=row_num, column=6).value = row_data["decision"]
        ws.cell(row=row_num, column=7).value = row_data["notes"]

        # Apply decision color
        decision_cell = ws.cell(row=row_num, column=6)
        decision = row_data["decision"]
        if decision in decision_colors:
            decision_cell.fill = decision_colors[decision]

        # Apply borders and alignment to all cells
        for col_num in range(1, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = center_alignment

    # Set column widths
    column_widths = {
        'A': 20,  # Legacy Page
        'B': 15,  # Design Number
        'C': 35,  # Boat Name
        'D': 35,  # New Site Page
        'E': 18,  # New Site Status
        'F': 18,  # Decision
        'G': 40,  # Notes
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze panes (freeze top row)
    ws.freeze_panes = "A2"

    # Generate summary statistics
    summary_row = len(review_rows) + 3
    stats = defaultdict(int)
    for row in review_rows:
        stats[row["decision"]] += 1
    stats["Total"] = len(review_rows)

    ws.cell(row=summary_row, column=1).value = "SUMMARY"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=11)

    for idx, (decision, count) in enumerate(sorted(stats.items()), 0):
        ws.cell(row=summary_row + 1 + idx, column=1).value = f"{decision}:"
        ws.cell(row=summary_row + 1 + idx, column=2).value = count
        ws.cell(row=summary_row + 1 + idx, column=1).font = Font(bold=True)

    # Save
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"\nSpreadsheet saved to: {OUTPUT_PATH}")

    # Print summary
    print("\n" + "=" * 60)
    print("MIGRATION SUMMARY")
    print("=" * 60)
    for decision, count in sorted(stats.items()):
        print(f"{decision:.<40} {count:>6}")
    print("=" * 60)


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("\n" + "=" * 60)
    print("ARCHIVE PARITY REVIEW GENERATOR")
    print("=" * 60 + "\n")

    # Scan both sites
    old_pages = scan_old_site()
    new_pages = scan_new_site()

    # Generate review data
    print("\nMatching and classifying pages...")
    review_rows = generate_review_data(old_pages, new_pages)

    # Create spreadsheet
    print("Creating spreadsheet...")
    create_spreadsheet(review_rows)

    print("\n✓ Complete!")


if __name__ == "__main__":
    main()
